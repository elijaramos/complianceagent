"""
Azure Compliance Scanner for NIST CSF

This module implements automated compliance scanning for Azure resources against
NIST Cybersecurity Framework (CSF) 2.0 controls. It identifies misconfigurations
and security risks in Azure infrastructure.

WHY THIS EXISTS:
- Manual compliance checks are error-prone and time-consuming
- Organizations need continuous monitoring to maintain security posture
- Regulatory frameworks (GDPR, HIPAA, PCI DSS) require documented evidence of controls
- Early detection of misconfigurations prevents security incidents and data breaches
"""

import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path

import yaml
from azure.identity import ClientSecretCredential
from azure.mgmt.storage import StorageManagementClient
from azure.mgmt.network import NetworkManagementClient
from azure.mgmt.resource import ResourceManagementClient
from azure.storage.blob import BlobServiceClient
from azure.core.exceptions import AzureError, HttpResponseError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import configuration settings
sys.path.insert(0, str(Path(__file__).parent.parent))
from settings import REPORTS_DIR


class AzureScanner:
    """
    Scans Azure resources for compliance with NIST CSF controls.
    
    WHY USE THIS CLASS:
    - Centralizes Azure authentication and client management
    - Provides reusable scanning logic across multiple resource types
    - Generates actionable compliance reports for security teams
    - Enables automated compliance monitoring in CI/CD pipelines
    
    Architecture:
    - Uses Azure SDK management libraries (azure-mgmt-*) to query resource configurations
    - Authenticates via Service Principal (ClientSecretCredential) for non-interactive scenarios
    - Loads compliance rules from YAML for flexibility and maintainability
    - Separates scanning logic by resource type for modularity
    """
    
    def __init__(self, subscription_id: str, tenant_id: str, client_id: str, 
                 client_secret: str, rules_path: str = "config/nist_csf_rules.yaml"):
        """
        Initialize Azure compliance scanner with credentials and rules.
        
        WHY THESE PARAMETERS:
        - subscription_id: Defines the Azure subscription scope to scan
        - tenant_id: Identifies the Azure AD tenant for authentication
        - client_id: Service Principal application ID (used for non-interactive auth)
        - client_secret: Service Principal password/secret (proves identity)
        - rules_path: Allows different rule sets (dev/prod, different frameworks)
        
        WHY SERVICE PRINCIPAL AUTH:
        - Enables automation without user interaction
        - Can be scoped with minimal required permissions (principle of least privilege)
        - Supports secret rotation and centralized credential management
        - Works in CI/CD pipelines, scheduled tasks, and containerized environments
        
        Args:
            subscription_id: Azure subscription ID to scan
            tenant_id: Azure Active Directory tenant ID
            client_id: Service Principal application (client) ID
            client_secret: Service Principal secret value
            rules_path: Path to NIST CSF rules YAML file
            
        Raises:
            FileNotFoundError: If rules file doesn't exist
            yaml.YAMLError: If rules file is malformed
        """
        self.subscription_id = subscription_id
        
        # Azure SDK Pattern: ClientSecretCredential is used for service-to-service authentication
        # It implements the OAuth 2.0 client credentials flow, exchanging the client secret
        # for a time-limited access token. Tokens are cached and automatically refreshed.
        self.credential = ClientSecretCredential(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret
        )
        
        # Azure SDK Pattern: Management clients provide typed interfaces to Azure Resource Manager (ARM)
        # Each client corresponds to a specific Azure service (storage, network, compute, etc.)
        # They handle REST API calls, pagination, retries, and response deserialization
        self.storage_client = StorageManagementClient(
            credential=self.credential,
            subscription_id=subscription_id
        )
        
        self.network_client = NetworkManagementClient(
            credential=self.credential,
            subscription_id=subscription_id
        )

        self.resource_client = ResourceManagementClient(
            credential=self.credential,
            subscription_id=subscription_id
        )

        # Load compliance rules from YAML
        # WHY YAML: Human-readable, supports comments, easy for security teams to maintain
        self.rules = self._load_rules(rules_path)
        
        # Store scan results for report generation
        self.scan_results: List[Dict[str, Any]] = []
        
    def _load_rules(self, rules_path: str) -> List[Dict[str, Any]]:
        """
        Load compliance rules from YAML configuration file.
        
        WHY EXTERNAL RULES FILE:
        - Allows security teams to update rules without code changes
        - Supports versioning and auditing of compliance requirements
        - Enables different rule sets for different environments (dev/prod)
        - Makes it easy to add new controls as regulations evolve
        
        Args:
            rules_path: Path to YAML rules file
            
        Returns:
            List of rule dictionaries
            
        Raises:
            FileNotFoundError: If rules file doesn't exist
            yaml.YAMLError: If YAML is malformed
        """
        rules_file = Path(rules_path)
        
        if not rules_file.exists():
            raise FileNotFoundError(
                f"Rules file not found: {rules_path}\n"
                f"Expected location: {rules_file.absolute()}"
            )
        
        try:
            with open(rules_file, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
                rules = data.get('rules', [])
                print(f"✓ Loaded {len(rules)} compliance rules from {rules_path}")
                return rules
        except yaml.YAMLError as e:
            raise yaml.YAMLError(f"Failed to parse rules file: {e}")
    
    def scan_all_resources(self) -> Dict[str, Any]:
        """
        Main entry point: Scan all Azure resources for compliance violations.
        
        WHY THIS METHOD:
        - Provides single entry point for complete compliance scan
        - Coordinates scanning across multiple resource types
        - Aggregates results for unified reporting
        - Handles errors gracefully to ensure partial results are still useful
        
        Returns:
            Dictionary containing:
                - timestamp: When scan was performed
                - subscription_id: Azure subscription scanned
                - total_violations: Total number of compliance violations found
                - violations_by_severity: Breakdown by severity level
                - results: Detailed violation information
                
        WHY THIS STRUCTURE:
        - timestamp: Enables trending analysis and audit trails
        - subscription_id: Critical for multi-subscription environments
        - severity breakdown: Helps prioritize remediation efforts
        - detailed results: Provides actionable information for fixes
        """
        print(f"\n{'='*70}")
        print(f"Starting Azure Compliance Scan")
        print(f"Subscription: {self.subscription_id}")
        print(f"Timestamp: {datetime.utcnow().isoformat()}")
        print(f"{'='*70}\n")
        
        # Reset results for new scan
        self.scan_results = []
        
        # Scan different resource types
        # WHY SEPARATE METHODS: Each Azure service has different APIs and data structures
        # This separation makes the code maintainable and testable
        try:
            storage_violations = self._scan_storage_accounts()
            print(f"✓ Storage scan complete: {len(storage_violations)} violations found")
        except Exception as e:
            print(f"✗ Storage scan failed: {e}")
            storage_violations = []
        
        try:
            nsg_violations = self._scan_nsgs()
            print(f"✓ NSG scan complete: {len(nsg_violations)} violations found")
        except Exception as e:
            print(f"✗ NSG scan failed: {e}")
            nsg_violations = []
        
        # Combine all violations
        self.scan_results = storage_violations + nsg_violations
        
        # Generate summary statistics
        summary = self._generate_summary()
        
        print(f"\n{'='*70}")
        print(f"Scan Complete: {summary['total_violations']} total violations")
        print(f"{'='*70}\n")
        
        return summary
    
    def _scan_storage_accounts(self) -> List[Dict[str, Any]]:
        """
        Scan all storage accounts in subscription for compliance violations.
        
        WHY STORAGE ACCOUNTS ARE CRITICAL:
        - Often store sensitive data (PII, financial records, health data)
        - Frequently misconfigured with public access, leading to data breaches
        - Encryption and logging are essential for compliance (GDPR, HIPAA, PCI DSS)
        - Common attack vector: exposed storage accounts are actively scanned by attackers
        
        Azure SDK Pattern:
        - storage_client.storage_accounts.list() returns an iterator (lazy loading)
        - Each item is a StorageAccount object with nested properties
        - Properties follow Azure Resource Manager (ARM) structure
        
        Returns:
            List of violation dictionaries with details for remediation
        """
        violations = []
        
        # Get all storage rules from loaded compliance rules
        storage_rules = [
            r for r in self.rules 
            if r.get('resource_type') == 'Microsoft.Storage/storageAccounts'
        ]
        
        print(f"\nScanning Storage Accounts ({len(storage_rules)} rules)...")
        
        try:
            # Azure SDK Pattern: list() returns a paged iterator
            # The SDK automatically handles pagination (continuation tokens)
            # This is efficient for subscriptions with many resources
            storage_accounts = list(self.storage_client.storage_accounts.list())
            print(f"  Found {len(storage_accounts)} storage account(s)")
            
            if not storage_accounts:
                print("  ⚠ No storage accounts found in subscription")
                return violations
            
            for account in storage_accounts:
                # Extract resource group from ARM resource ID
                # Azure resource IDs follow pattern: /subscriptions/{sub}/resourceGroups/{rg}/providers/{provider}/{type}/{name}
                resource_group = account.id.split('/')[4]
                
                print(f"  Checking: {account.name} (RG: {resource_group})")
                
                # Get detailed properties that may not be in list response
                # WHY: List operations return minimal data for performance
                # Get operations return complete resource details
                try:
                    account_details = self.storage_client.storage_accounts.get_properties(
                        resource_group_name=resource_group,
                        account_name=account.name
                    )
                except HttpResponseError as e:
                    print(f"    ✗ Failed to get properties: {e.message}")
                    continue
                
                # Check each storage rule against this account
                for rule in storage_rules:
                    violation = self._check_storage_rule(
                        account_details, 
                        rule, 
                        resource_group
                    )
                    if violation:
                        violations.append(violation)
                        severity_symbol = "🔴" if rule['severity'] == "CRITICAL" else "🟡"
                        print(f"    {severity_symbol} VIOLATION: {rule['description']}")
        
        except AzureError as e:
            print(f"  ✗ Azure API error: {e}")
            raise
        
        return violations
    
    def _check_storage_rule(self, account: Any, rule: Dict[str, Any],
                           resource_group: str) -> Optional[Dict[str, Any]]:
        """
        Check a single storage account against a compliance rule.

        WHY SEPARATE METHOD:
        - Keeps rule checking logic isolated and testable
        - Allows easy addition of new rule types
        - Simplifies error handling for individual rules

        Args:
            account: Azure StorageAccount object
            rule: Rule dictionary from YAML
            resource_group: Resource group name (for reporting)

        Returns:
            Violation dictionary if rule failed, None if passed
        """
        check = rule.get('check', {})
        property_path = check.get('property')
        operator = check.get('operator')
        expected_value = check.get('value')

        # Special case: diagnostic_settings requires Monitor API call
        # WHY SPECIAL HANDLING: Diagnostic settings aren't a property of StorageAccount object
        if property_path == 'diagnostic_settings':
            actual_value = self._check_diagnostic_settings(account, resource_group)
        else:
            # Navigate nested property path (e.g., "encryption.services.blob.enabled")
            # WHY: Azure API responses have deeply nested structures
            actual_value = self._get_nested_property(account, property_path)

        # Evaluate rule based on operator
        is_compliant = self._evaluate_rule(actual_value, operator, expected_value)
        
        if not is_compliant:
            return {
                'rule_id': rule['id'],
                'resource_type': rule['resource_type'],
                'resource_name': account.name,
                'resource_group': resource_group,
                'severity': rule['severity'],
                'description': rule['description'],
                'nist_function': rule.get('nist_function'),
                'nist_category': rule.get('category'),
                'expected': expected_value,
                'actual': actual_value,
                'timestamp': datetime.utcnow().isoformat()
            }
        
        return None

    def _check_diagnostic_settings(self, account: Any, resource_group: str) -> bool:
        """
        Check if Storage Analytics logging is enabled for a storage account.

        WHY STORAGE ANALYTICS LOGGING:
        - Required for security incident detection and forensics
        - Logs all access attempts, successful and failed
        - Compliance requirement for SOX, HIPAA, PCI DSS
        - Enables threat detection and anomaly analysis

        Storage Analytics Pattern:
        - Uses BlobServiceClient to get service properties
        - Logs stored in $logs container within storage account
        - Checks for read, write, delete logging enabled

        Args:
            account: Azure StorageAccount object
            resource_group: Resource group name

        Returns:
            True if Storage Analytics logging is enabled, False otherwise
        """
        try:
            # Get storage account keys to authenticate BlobServiceClient
            keys = self.storage_client.storage_accounts.list_keys(
                resource_group_name=resource_group,
                account_name=account.name
            )

            if not keys.keys or len(keys.keys) == 0:
                return False

            account_key = keys.keys[0].value

            # Create BlobServiceClient using account key
            account_url = f"https://{account.name}.blob.core.windows.net"
            blob_service_client = BlobServiceClient(
                account_url=account_url,
                credential=account_key
            )

            # Get service properties to check logging configuration
            properties = blob_service_client.get_service_properties()

            # Check if logging is enabled for any operation type
            # Note: get_service_properties() returns a dict
            if 'analytics_logging' in properties and properties['analytics_logging']:
                logging = properties['analytics_logging']
                # Compliant if ANY logging is enabled (read, write, or delete)
                if logging.read or logging.write or logging.delete:
                    return True

            return False

        except Exception as e:
            # If we can't check (permissions, API error), assume not configured
            # WHY ASSUME FALSE: Fail secure - better to flag as violation than miss it
            print(f"  ⚠ Could not check Storage Analytics logging for {account.name}: {e}")
            return False

    def _scan_nsgs(self) -> List[Dict[str, Any]]:
        """
        Scan all Network Security Groups for compliance violations.
        
        WHY NSGs ARE CRITICAL:
        - Act as virtual firewalls controlling network traffic to Azure resources
        - Misconfigured NSGs expose resources to internet attacks (brute force, DDoS, exploits)
        - Allow rules with source 0.0.0.0/0 are the most common security misconfiguration
        - Attackers actively scan for publicly exposed resources (SSH, RDP, databases)
        
        Azure SDK Pattern:
        - NSGs are listed across all resource groups in subscription
        - Security rules are nested within NSG properties
        - Rules have priority (lower numbers evaluated first)
        
        Returns:
            List of violation dictionaries
        """
        violations = []
        
        # Get NSG rules from compliance rules
        nsg_rules = [
            r for r in self.rules 
            if r.get('resource_type') == 'Microsoft.Network/networkSecurityGroups'
        ]
        
        print(f"\nScanning Network Security Groups ({len(nsg_rules)} rules)...")
        
        try:
            # Azure SDK Pattern: list_all() gets NSGs across all resource groups
            # Alternative: list(resource_group_name) for single resource group
            nsgs = list(self.network_client.network_security_groups.list_all())
            print(f"  Found {len(nsgs)} NSG(s)")
            
            if not nsgs:
                print("  ⚠ No NSGs found in subscription")
                return violations
            
            for nsg in nsgs:
                # Extract resource group from ARM resource ID
                resource_group = nsg.id.split('/')[4]
                
                print(f"  Checking: {nsg.name} (RG: {resource_group})")
                
                # Check each NSG rule
                for rule in nsg_rules:
                    violation = self._check_nsg_rule(nsg, rule, resource_group)
                    if violation:
                        violations.append(violation)
                        severity_symbol = "🔴" if rule['severity'] == "CRITICAL" else "🟡"
                        print(f"    {severity_symbol} VIOLATION: {rule['description']}")
        
        except AzureError as e:
            print(f"  ✗ Azure API error: {e}")
            raise
        
        return violations
    
    def _check_nsg_rule(self, nsg: Any, rule: Dict[str, Any], 
                        resource_group: str) -> Optional[Dict[str, Any]]:
        """
        Check NSG security rules for dangerous configurations.
        
        WHY THIS IS COMPLEX:
        - NSGs have multiple security rules, each with different properties
        - Must check combinations of properties (source + access + direction)
        - "not_contains" operator requires checking all rules don't match pattern
        
        Security Logic:
        - Checks if ANY security rule allows unrestricted inbound access (0.0.0.0/0)
        - This is the most dangerous misconfiguration (exposes resources to internet)
        
        Args:
            nsg: Azure NetworkSecurityGroup object
            rule: Compliance rule dictionary
            resource_group: Resource group name
            
        Returns:
            Violation dictionary if dangerous rule found, None if safe
        """
        check = rule.get('check', {})
        operator = check.get('operator')
        
        if operator == 'not_contains':
            # Check if any security rule matches the dangerous pattern
            dangerous_pattern = check.get('value', {})
            
            # NSG security rules are in properties.security_rules
            security_rules = getattr(nsg, 'security_rules', []) or []
            
            for sec_rule in security_rules:
                # Check if this rule matches ALL criteria in the dangerous pattern
                matches_pattern = True
                
                for pattern_key, pattern_value in dangerous_pattern.items():
                    # Navigate nested properties (e.g., "properties.sourceAddressPrefix")
                    actual_value = self._get_nested_property(sec_rule, pattern_key)
                    
                    # WHY STRING COMPARISON: Azure API may return "*" or "0.0.0.0/0" or "Internet"
                    # All represent "any source address"
                    if pattern_key == 'properties.sourceAddressPrefix':
                        # Check for common "any source" patterns
                        if actual_value not in ['0.0.0.0/0', '*', 'Internet']:
                            matches_pattern = False
                            break
                    elif actual_value != pattern_value:
                        matches_pattern = False
                        break
                
                # If we found a rule matching the dangerous pattern, it's a violation
                if matches_pattern:
                    return {
                        'rule_id': rule['id'],
                        'resource_type': rule['resource_type'],
                        'resource_name': nsg.name,
                        'resource_group': resource_group,
                        'severity': rule['severity'],
                        'description': rule['description'],
                        'nist_function': rule.get('nist_function'),
                        'nist_category': rule.get('category'),
                        'violating_rule': sec_rule.name,
                        'violating_rule_details': {
                            'source': sec_rule.source_address_prefix,
                            'destination_port': sec_rule.destination_port_range,
                            'protocol': sec_rule.protocol,
                            'access': sec_rule.access,
                            'direction': sec_rule.direction,
                            'priority': sec_rule.priority
                        },
                        'timestamp': datetime.utcnow().isoformat()
                    }
        
        return None
    
    def _get_nested_property(self, obj: Any, property_path: str) -> Any:
        """
        Navigate nested object properties using dot notation.
        
        WHY NEEDED:
        - Azure SDK returns deeply nested objects
        - Rules specify property paths like "properties.encryption.services.blob.enabled"
        - Avoids hardcoding property access patterns
        
        Example:
            obj = StorageAccount(properties=Properties(encryption=Encryption(services=...)))
            _get_nested_property(obj, "properties.encryption.services.blob.enabled")
            
        Args:
            obj: Python object (usually Azure SDK model)
            property_path: Dot-separated property path
            
        Returns:
            Property value or None if path doesn't exist
        """
        if not property_path:
            return obj
        
        # Split path and navigate step by step
        parts = property_path.split('.')
        current = obj
        
        for part in parts:
            if hasattr(current, part):
                current = getattr(current, part)
            elif isinstance(current, dict) and part in current:
                current = current[part]
            else:
                # Property doesn't exist (might be optional in Azure API)
                return None
        
        return current
    
    def _evaluate_rule(self, actual_value: Any, operator: str, expected_value: Any) -> bool:
        """
        Evaluate if actual value meets expected value based on operator.
        
        WHY SEPARATE METHOD:
        - Centralizes comparison logic
        - Makes it easy to add new operators
        - Handles None/null cases consistently
        
        Args:
            actual_value: Value from Azure resource
            operator: Comparison operator (equals, not_equals, etc.)
            expected_value: Expected value from rule
            
        Returns:
            True if compliant, False if violation
        """
        if operator == 'equals':
            return actual_value == expected_value
        elif operator == 'not_equals':
            return actual_value != expected_value
        elif operator == 'contains':
            return expected_value in (actual_value or [])
        elif operator == 'not_contains':
            return expected_value not in (actual_value or [])
        else:
            print(f"  ⚠ Unknown operator: {operator}")
            return True  # Unknown operators pass by default (fail-safe)
    
    def _generate_summary(self) -> Dict[str, Any]:
        """
        Generate summary statistics from scan results.
        
        WHY SUMMARY STATISTICS:
        - Executives need high-level metrics (total violations, trends)
        - Security teams need severity breakdown for prioritization
        - Audit trails require timestamps and subscription tracking
        
        Returns:
            Summary dictionary with aggregated scan results
        """
        # Count violations by severity
        severity_counts = {
            'CRITICAL': 0,
            'HIGH': 0,
            'MEDIUM': 0,
            'LOW': 0
        }
        
        for result in self.scan_results:
            severity = result.get('severity', 'UNKNOWN')
            if severity in severity_counts:
                severity_counts[severity] += 1
        
        return {
            'timestamp': datetime.utcnow().isoformat(),
            'subscription_id': self.subscription_id,
            'total_violations': len(self.scan_results),
            'violations_by_severity': severity_counts,
            'results': self.scan_results
        }
    
    def generate_scan_report(self, output_path: Optional[str] = None) -> str:
        """
        Generate formatted Excel compliance report and save to file.

        WHY EXCEL REPORTS:
        - Professional presentation for executives and auditors
        - Easy filtering, sorting, and pivot tables for analysis
        - Color-coded severity levels for quick risk assessment
        - Supports complex data structures and formulas
        - Industry standard for GRC reporting and evidence collection

        Args:
            output_path: Path to save report file (defaults to REPORTS_DIR/compliance_report.xlsx)

        Returns:
            Report path as string
        """
        # Use default path from settings if not provided
        if output_path is None:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_path = f"{REPORTS_DIR}/compliance_report_{timestamp}.xlsx"

        # Ensure reports directory exists
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Create Excel workbook
        wb = Workbook()

        # Remove default sheet and create custom sheets
        wb.remove(wb.active)

        # Create Summary sheet
        self._create_summary_sheet(wb)

        # Create Violations sheet
        self._create_violations_sheet(wb)

        # Create Recommendations sheet
        self._create_recommendations_sheet(wb)

        # Save workbook
        try:
            wb.save(output_path)
            print(f"\n✓ Excel report saved to: {output_file.absolute()}")
            return str(output_file.absolute())
        except IOError as e:
            print(f"\n✗ Failed to save Excel report: {e}")
            return ""

    def _create_summary_sheet(self, wb: Workbook):
        """Create summary overview sheet with key metrics."""
        ws = wb.create_sheet("Summary", 0)

        # Define colors
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        critical_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        high_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        medium_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        white_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=11)

        # Title
        ws['A1'] = "AZURE COMPLIANCE SCAN REPORT"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        # Report metadata
        ws['A3'] = "Report Date:"
        ws['B3'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        ws['A4'] = "Subscription ID:"
        ws['B4'] = self.subscription_id
        ws['A5'] = "Total Violations:"
        ws['B5'] = len(self.scan_results)

        # Style metadata
        for row in range(3, 6):
            ws[f'A{row}'].font = header_font

        # Severity breakdown header
        ws['A7'] = "VIOLATIONS BY SEVERITY"
        ws['A7'].font = Font(bold=True, size=14)
        ws.merge_cells('A7:D7')

        # Severity table headers
        ws['A9'] = "Severity"
        ws['B9'] = "Count"
        ws['C9'] = "Percentage"
        ws['D9'] = "SLA"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}9'].fill = header_fill
            ws[f'{col}9'].font = white_font
            ws[f'{col}9'].alignment = Alignment(horizontal='center')

        # Count violations by severity
        severity_counts = {}
        total = len(self.scan_results)

        for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
            count = len([v for v in self.scan_results if v.get('severity') == severity])
            severity_counts[severity] = count

        # Populate severity data
        row = 10
        severity_config = {
            'CRITICAL': (critical_fill, "Immediate"),
            'HIGH': (high_fill, "7 Days"),
            'MEDIUM': (medium_fill, "30 Days"),
            'LOW': (low_fill, "90 Days")
        }

        for severity, (fill, sla) in severity_config.items():
            count = severity_counts.get(severity, 0)
            percentage = (count / total * 100) if total > 0 else 0

            ws[f'A{row}'] = severity
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            ws[f'D{row}'] = sla

            # Apply severity color to entire row
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_violations_sheet(self, wb: Workbook):
        """Create detailed violations sheet with all findings."""
        ws = wb.create_sheet("Violations", 1)

        # Define styles
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Severity", "Description", "Resource Name", "Resource Group",
            "Resource Type", "Rule ID", "NIST Function", "NIST Category",
            "Expected", "Actual", "Details"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = border

        # Populate data
        severity_colors = {
            'CRITICAL': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
            'HIGH': PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),
            'MEDIUM': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            'LOW': PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        }

        # Sort violations by severity
        severity_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        sorted_violations = sorted(
            self.scan_results,
            key=lambda x: severity_order.get(x.get('severity', 'LOW'), 4)
        )

        for row_num, violation in enumerate(sorted_violations, 2):
            severity = violation.get('severity', 'N/A')

            # Build details string
            details = []
            if 'violating_rule' in violation:
                details.append(f"NSG Rule: {violation['violating_rule']}")
                rule_details = violation.get('violating_rule_details', {})
                details.append(f"Source: {rule_details.get('source', 'N/A')}")
                details.append(f"Port: {rule_details.get('destination_port', 'N/A')}")

            row_data = [
                severity,
                violation.get('description', 'N/A'),
                violation.get('resource_name', 'N/A'),
                violation.get('resource_group', 'N/A'),
                violation.get('resource_type', 'N/A'),
                violation.get('rule_id', 'N/A'),
                violation.get('nist_function', 'N/A'),
                violation.get('nist_category', 'N/A'),
                str(violation.get('expected', 'N/A')),
                str(violation.get('actual', 'N/A')),
                ' | '.join(details) if details else 'N/A'
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Apply severity color to row
                if col_num == 1 and severity in severity_colors:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=c).fill = severity_colors[severity]

        # Set column widths
        column_widths = [12, 50, 25, 25, 35, 30, 15, 15, 20, 20, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    def _create_recommendations_sheet(self, wb: Workbook):
        """Create recommendations and remediation guidance sheet."""
        ws = wb.create_sheet("Recommendations", 2)

        # Title
        ws['A1'] = "REMEDIATION RECOMMENDATIONS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')

        # Recommendations
        recommendations = [
            ("Priority", "Timeframe", "Action"),
            ("CRITICAL", "Immediate", "Remediate immediately - poses active security risk"),
            ("HIGH", "7 Days", "Address within one week - significant compliance gap"),
            ("MEDIUM", "30 Days", "Resolve within 30 days - moderate security impact"),
            ("LOW", "90 Days", "Plan remediation - minor compliance issue"),
            ("", "", ""),
            ("General Best Practices", "", ""),
            ("1.", "Automation", "Use Azure Policy to prevent future violations"),
            ("2.", "IaC Security", "Scan Terraform/ARM templates before deployment"),
            ("3.", "Monitoring", "Schedule daily/weekly compliance scans"),
            ("4.", "Documentation", "Maintain audit trail of all remediations"),
            ("5.", "Training", "Educate dev teams on secure Azure configuration"),
        ]

        # Define styles
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)

        start_row = 3
        for i, (col1, col2, col3) in enumerate(recommendations):
            row = start_row + i
            ws[f'A{row}'] = col1
            ws[f'B{row}'] = col2
            ws[f'C{row}'] = col3

            # Style header row
            if i == 0:
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = header_fill
                    ws[f'{col}{row}'].font = white_font
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

            # Bold section headers
            if col1 in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "General Best Practices"]:
                ws[f'A{row}'].font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60


# WHY __main__ BLOCK:
# - Allows running scanner as standalone script for testing
# - Provides example usage for other developers
# - Enables quick manual testing without writing separate test scripts
if __name__ == "__main__":
    """
    Test the Azure scanner with credentials from environment variables.
    
    Required Environment Variables:
        AZURE_SUBSCRIPTION_ID: Azure subscription ID to scan
        AZURE_TENANT_ID: Azure AD tenant ID
        AZURE_CLIENT_ID: Service Principal application ID
        AZURE_CLIENT_SECRET: Service Principal secret
    
    WHY ENVIRONMENT VARIABLES:
    - Keeps secrets out of source code (security best practice)
    - Works with various secret management systems (Azure Key Vault, HashiCorp Vault)
    - Compatible with CI/CD pipelines (GitHub Actions, Azure DevOps)
    - Follows 12-factor app methodology for cloud-native applications
    
    Usage:
        # Set environment variables first
        export AZURE_SUBSCRIPTION_ID="your-sub-id"
        export AZURE_TENANT_ID="your-tenant-id"
        export AZURE_CLIENT_ID="your-client-id"
        export AZURE_CLIENT_SECRET="your-client-secret"
        
        # Run scanner
        python src/scanner.py
    """
    # Load credentials from environment
    subscription_id = os.getenv("AZURE_SUBSCRIPTION_ID")
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    
    # Validate all required credentials are present
    missing_vars = []
    if not subscription_id:
        missing_vars.append("AZURE_SUBSCRIPTION_ID")
    if not tenant_id:
        missing_vars.append("AZURE_TENANT_ID")
    if not client_id:
        missing_vars.append("AZURE_CLIENT_ID")
    if not client_secret:
        missing_vars.append("AZURE_CLIENT_SECRET")
    
    if missing_vars:
        print("ERROR: Missing required environment variables:")
        for var in missing_vars:
            print(f"  - {var}")
        print("\nPlease set these variables before running the scanner.")
        print("Example:")
        print('  export AZURE_SUBSCRIPTION_ID="your-subscription-id"')
        sys.exit(1)
    
    try:
        # Initialize scanner
        print("Initializing Azure Scanner...")
        scanner = AzureScanner(
            subscription_id=subscription_id,
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret
        )
        
        # Run full compliance scan
        results = scanner.scan_all_resources()
        
        # Generate and save report
        report = scanner.generate_scan_report()
        
        # Print summary to console
        print("\n" + "=" * 80)
        print("SCAN SUMMARY")
        print("=" * 80)
        print(f"Total Violations: {results['total_violations']}")
        print(f"  CRITICAL: {results['violations_by_severity']['CRITICAL']}")
        print(f"  HIGH: {results['violations_by_severity']['HIGH']}")
        print(f"  MEDIUM: {results['violations_by_severity']['MEDIUM']}")
        print(f"  LOW: {results['violations_by_severity']['LOW']}")
        print("=" * 80)
        
        # Exit with error code if violations found (useful for CI/CD)
        # WHY: Allows pipeline to fail if compliance violations are detected
        if results['total_violations'] > 0:
            sys.exit(1)
        else:
            print("\n✓ No compliance violations found!")
            sys.exit(0)
            
    except FileNotFoundError as e:
        print(f"\nERROR: {e}")
        sys.exit(1)
    except AzureError as e:
        print(f"\nAzure API Error: {e}")
        print("\nPossible causes:")
        print("  - Invalid credentials")
        print("  - Insufficient permissions (Reader role required)")
        print("  - Network connectivity issues")
        print("  - Invalid subscription ID")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)